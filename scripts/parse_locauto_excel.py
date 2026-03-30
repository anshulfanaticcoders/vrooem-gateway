#!/usr/bin/env python3
"""Generate Locauto location YAML from the official Excel workbook."""

from __future__ import annotations

import argparse
import re
from pathlib import Path

import openpyxl
import yaml


def _titleize(value: str | None) -> str:
    if not value:
        return ""
    return " ".join(part.capitalize() for part in str(value).strip().split())


def _normalize_text(value: str | None) -> str | None:
    if value is None:
        return None
    cleaned = " ".join(str(value).split())
    return cleaned or None


def _normalize_phone(value: str | None) -> str | None:
    cleaned = _normalize_text(value)
    if not cleaned:
        return None
    return cleaned if cleaned.startswith("+39") else f"+39 {cleaned}"


def _normalize_hours(value: str | None) -> str | None:
    cleaned = _normalize_text(value)
    if not cleaned:
        return None

    def repl(match: re.Match[str]) -> str:
        return f"{int(match.group(1)):02d}:{int(match.group(2)):02d}"

    return re.sub(r"(\d{1,2})\.(\d{2})", repl, cleaned)


def _parse_coords(value: str | None) -> tuple[float | None, float | None]:
    cleaned = _normalize_text(value)
    if not cleaned or "/" not in cleaned:
        return None, None
    left, right = [segment.strip() for segment in cleaned.split("/", 1)]
    try:
        return float(left), float(right)
    except ValueError:
        return None, None


def _postal_code(address: str | None) -> str | None:
    if not address:
        return None
    match = re.search(r"\b(\d{5})\b", address)
    return match.group(1) if match else None


def _infer_location_type(name: str) -> str:
    lowered = name.lower()
    if "airport" in lowered:
        return "airport"
    if "station" in lowered:
        return "train_station"
    return "downtown"


def _infer_iata(code: str, location_type: str) -> str | None:
    normalized = code.strip().upper()
    if location_type != "airport":
        return None
    return normalized if len(normalized) == 3 and normalized.isalpha() else None


def _sheet_index(workbook, sheet_name: str, code_index: int = 1) -> dict[str, tuple]:
    sheet = workbook[sheet_name]
    indexed: dict[str, tuple] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        code = row[code_index]
        if code:
            indexed[str(code).strip().upper()] = row
    return indexed


def build_locations(workbook_path: Path) -> list[dict]:
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)

    details = _sheet_index(workbook, "Location details")
    hours = _sheet_index(workbook, "Standard opening hours")
    holidays_sheet = workbook["Bank hols 2026"]
    holiday_headers = [
        _normalize_text(value) or f"holiday_{index}"
        for index, value in enumerate(next(holidays_sheet.iter_rows(min_row=1, max_row=1, values_only=True))[2:], start=1)
    ]
    holidays = _sheet_index(workbook, "Bank hols 2026")
    out_of_hours = _sheet_index(workbook, "OOH rules ")
    wayfinding = _sheet_index(workbook, "Wayfinding")

    locations: list[dict] = []

    for code, row in sorted(details.items()):
        raw_name = _normalize_text(row[0]) or code
        name = _titleize(raw_name)
        city = _titleize(row[3])
        province = _titleize(row[4])
        latitude, longitude = _parse_coords(row[5])
        address = _normalize_text(row[2])
        location_type = _infer_location_type(name)

        location_hours = hours.get(code, ())
        holiday_row = holidays.get(code, ())
        ooh_row = out_of_hours.get(code, ())
        wayfinding_row = wayfinding.get(code, ())

        holiday_hours = {}
        for header, value in zip(holiday_headers, holiday_row[2:]):
            normalized_holiday_hours = _normalize_hours(value)
            if normalized_holiday_hours:
                holiday_hours[header] = normalized_holiday_hours

        locations.append(
            {
                "code": code,
                "name": name,
                "city": city,
                "province": province,
                "country": "Italy",
                "country_code": "IT",
                "location_type": location_type,
                "is_airport": location_type == "airport",
                "iata": _infer_iata(code, location_type),
                "latitude": latitude,
                "longitude": longitude,
                "address": address,
                "postal_code": _postal_code(address),
                "phone": _normalize_phone(row[7]),
                "email": _normalize_text(row[6]),
                "operating_hours": {
                    "weekday": _normalize_hours(location_hours[2] if len(location_hours) > 2 else None),
                    "saturday": _normalize_hours(location_hours[3] if len(location_hours) > 3 else None),
                    "sunday": _normalize_hours(location_hours[4] if len(location_hours) > 4 else None),
                },
                "holiday_hours": holiday_hours,
                "pickup_instructions": _normalize_text(wayfinding_row[2] if len(wayfinding_row) > 2 else None),
                "dropoff_instructions": _normalize_text(wayfinding_row[3] if len(wayfinding_row) > 3 else None),
                "out_of_hours": {
                    "key_box": _normalize_text(ooh_row[2] if len(ooh_row) > 2 else None),
                    "pickup": _normalize_text(ooh_row[3] if len(ooh_row) > 3 else None),
                    "exceptions": _normalize_text(ooh_row[4] if len(ooh_row) > 4 else None),
                },
            }
        )

    return locations


def main() -> None:
    gateway_root = Path(__file__).resolve().parents[1]
    default_workbook = gateway_root.parent / "carrental" / "locauto docs" / "LOCATION LIST 2026 (TO) 2.12.25.xlsx"
    default_output = gateway_root / "config" / "suppliers" / "locauto_rent_locations.yaml"

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=default_workbook)
    parser.add_argument("--output", type=Path, default=default_output)
    args = parser.parse_args()

    locations = build_locations(args.workbook)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", encoding="utf-8") as handle:
        yaml.safe_dump({"locations": locations}, handle, sort_keys=False, allow_unicode=True)

    print(f"Wrote {len(locations)} Locauto locations to {args.output}")


if __name__ == "__main__":
    main()
